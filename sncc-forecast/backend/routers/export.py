from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
from dateutil.relativedelta import relativedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import models
from database import get_db
from services.summary_engine import calculate_summary
from services.yield_engine import calculate_yield
from services.draw_analyzer import analyze_draws
from services.lb_engine import calculate_lb
from services.hhh_engine import calculate_hhh

router = APIRouter()

CURRENCY_FORMAT = '$#,##0'
PERCENT_FORMAT = '0.00%'
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")


def style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def currency_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value or 0)
    cell.number_format = CURRENCY_FORMAT
    return cell


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value)) if cell.value else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)


@router.get("/versions/{version_id}/export")
def export_version(version_id: int, db: Session = Depends(get_db)):
    version = db.query(models.ForecastVersion).filter(models.ForecastVersion.id == version_id).first()
    if not version:
        raise HTTPException(status_code=404, detail="Version not found")

    today = date.today()
    start_month = today.replace(day=1)
    num_months = 12
    months = [(start_month + relativedelta(months=m)) for m in range(num_months)]
    month_labels = [m.strftime("%b %Y") for m in months]

    try:
        summary = calculate_summary(version_id, db, num_months)
        yield_data = calculate_yield(version_id, db, num_months=num_months)
    except Exception:
        summary = None
        yield_data = None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- Summary Sheet ----
    ws_summary = wb.create_sheet("Summary")
    headers = ["Category"] + month_labels
    for col_idx, h in enumerate(headers, start=1):
        ws_summary.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_summary, 1, len(headers))

    rows_data = []
    if summary:
        rows_data = [
            ("SFR", summary.get("sfr", [])),
            ("Multifamily", summary.get("mf", [])),
            ("A&D", summary.get("ad", [])),
            ("Raw Land", summary.get("raw_land", [])),
            ("Finished Lots", summary.get("finished_lots", [])),
            ("Total Loans", summary.get("total_loans", [])),
            ("Land Bucket", summary.get("land_bucket", [])),
            ("HHH Investments", summary.get("hhh", [])),
            ("Total ALL", summary.get("total_all", [])),
            ("Variance", summary.get("variance", [])),
        ]

    for row_idx, (label, values) in enumerate(rows_data, start=2):
        ws_summary.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        for col_idx, val in enumerate(values, start=2):
            cell = currency_cell(ws_summary, row_idx, col_idx, val)
            if label == "Variance" and val:
                cell.font = Font(color="006400" if val >= 0 else "FF0000")
        if row_idx % 2 == 1:
            for col in range(1, len(headers) + 1):
                ws_summary.cell(row=row_idx, column=col).fill = ALT_FILL

    auto_width(ws_summary)

    # ---- Yield Sheet ----
    ws_yield = wb.create_sheet("Yield")
    yield_headers = ["Category"] + month_labels
    for col_idx, h in enumerate(yield_headers, start=1):
        ws_yield.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_yield, 1, len(yield_headers))

    if yield_data:
        yield_rows = [
            ("Loan Yield", yield_data.get("loan_yield", [])),
            ("Land Bucket Yield", yield_data.get("lb_yield", [])),
            ("Profit Sharing", yield_data.get("profit_sharing", [])),
            ("Total Income", yield_data.get("total_income", [])),
        ]
        for row_idx, (label, values) in enumerate(yield_rows, start=2):
            ws_yield.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            for col_idx, val in enumerate(values, start=2):
                currency_cell(ws_yield, row_idx, col_idx, val)

    auto_width(ws_yield)

    # ---- Loan Type Sheets ----
    loan_types = [
        ("SFR", "SFR"),
        ("MF", "MF"),
        ("A&D", "AD"),
        ("Raw Land", "RawLand"),
        ("Finished Lots", "FinishedLots"),
    ]

    all_loans = db.query(models.Loan).filter(models.Loan.version_id == version_id).all()

    for sheet_name, loan_type_code in loan_types:
        ws = wb.create_sheet(sheet_name)
        loans_of_type = [l for l in all_loans if l.loan_type == loan_type_code]

        loan_headers = ["Loan#", "Borrower", "Development", "Balance", "Remaining", "Rate", "Maturity", "Monthly Draw", "Flagged"] + month_labels
        for col_idx, h in enumerate(loan_headers, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        style_header_row(ws, 1, len(loan_headers))

        if loans_of_type:
            draws = analyze_draws(loans_of_type, num_months)
            draw_map = {v["loan_number"]: v for v in draws.values()}

            for row_idx, loan in enumerate(loans_of_type, start=2):
                ws.cell(row=row_idx, column=1, value=loan.loan_number)
                ws.cell(row=row_idx, column=2, value=loan.borrower)
                ws.cell(row=row_idx, column=3, value=loan.development_name)
                currency_cell(ws, row_idx, 4, loan.current_balance)
                currency_cell(ws, row_idx, 5, loan.loan_remaining)
                ws.cell(row=row_idx, column=6, value=loan.interest_rate).number_format = PERCENT_FORMAT
                ws.cell(row=row_idx, column=7, value=loan.maturity_date)
                currency_cell(ws, row_idx, 8, loan.monthly_draw_increase)
                ws.cell(row=row_idx, column=9, value="Yes" if loan.draw_flagged else "No")
                if loan.draw_flagged:
                    for col in range(1, 10):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

                loan_draws = draw_map.get(loan.loan_number, {})
                balances = loan_draws.get("balances", [0.0] * num_months)
                for m_idx, bal in enumerate(balances):
                    currency_cell(ws, row_idx, 10 + m_idx, bal)

        auto_width(ws)

    # ---- NHC Forecast Sheet ----
    ws_nhc = wb.create_sheet("New Home Construction Forecast")
    nhc_headers = ["Development", "Builder", "Type"] + month_labels
    for col_idx, h in enumerate(nhc_headers, start=1):
        ws_nhc.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_nhc, 1, len(nhc_headers))

    nhc_assumptions = db.query(models.NHCAssumption).filter(
        models.NHCAssumption.version_id == version_id
    ).all()

    developments = {}
    for a in nhc_assumptions:
        key = (a.development_name, a.builder or "")
        if key not in developments:
            developments[key] = {}
        fm = a.forecast_month
        if hasattr(fm, 'date'):
            fm = fm.date()
        developments[key][fm.isoformat()] = a

    row_idx = 2
    for (dev_name, builder), month_data in developments.items():
        for data_type in ["SF Starts", "MF Starts", "SF Payoffs", "MF Payoffs"]:
            ws_nhc.cell(row=row_idx, column=1, value=dev_name)
            ws_nhc.cell(row=row_idx, column=2, value=builder)
            ws_nhc.cell(row=row_idx, column=3, value=data_type)
            for m_idx, month in enumerate(months):
                a = month_data.get(month.isoformat())
                val = 0
                if a:
                    if data_type == "SF Starts":
                        val = a.sf_starts or 0
                    elif data_type == "MF Starts":
                        val = a.mf_starts or 0
                    elif data_type == "SF Payoffs":
                        val = a.sf_payoffs or 0
                    elif data_type == "MF Payoffs":
                        val = a.mf_payoffs or 0
                ws_nhc.cell(row=row_idx, column=4 + m_idx, value=val)
            row_idx += 1

    auto_width(ws_nhc)

    # ---- Profit Sharing Sheet ----
    ws_ps = wb.create_sheet("Profit Sharing")
    ps_headers = ["Builder", "Product Type", "Avg Profit/Unit"] + month_labels
    for col_idx, h in enumerate(ps_headers, start=1):
        ws_ps.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_ps, 1, len(ps_headers))

    ps_assumptions = db.query(models.ProfitSharingAssumption).filter(
        models.ProfitSharingAssumption.version_id == version_id
    ).all()

    builder_ps = {}
    for ps in ps_assumptions:
        key = (ps.builder, ps.product_type or "")
        if key not in builder_ps:
            builder_ps[key] = {"avg_profit": ps.avg_profit_per_unit, "monthly": {}}
        fm = ps.forecast_month
        if hasattr(fm, 'date'):
            fm = fm.date()
        builder_ps[key]["monthly"][fm.isoformat()] = ps.closings_count

    row_idx = 2
    for (builder, ptype), data in builder_ps.items():
        ws_ps.cell(row=row_idx, column=1, value=builder)
        ws_ps.cell(row=row_idx, column=2, value=ptype)
        currency_cell(ws_ps, row_idx, 3, data["avg_profit"])
        for m_idx, month in enumerate(months):
            closings = data["monthly"].get(month.isoformat(), 0)
            ws_ps.cell(row=row_idx, column=4 + m_idx, value=closings)
        row_idx += 1

    auto_width(ws_ps)

    # ---- Land Bucket Sheet ----
    ws_lb = wb.create_sheet("LB Forecast")
    lb_headers = ["Dept Code", "Builder", "Project", "Phase", "Current Balance", "Total Budget"] + month_labels
    for col_idx, h in enumerate(lb_headers, start=1):
        ws_lb.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_lb, 1, len(lb_headers))

    lb_projects = db.query(models.LBProject).filter(
        models.LBProject.version_id == version_id
    ).all()

    lb_result = None
    if lb_projects:
        monthly_assumptions_raw = {}
        for proj in lb_projects:
            monthly_assumptions_raw[proj.id] = {}
            for ma in proj.monthly_assumptions:
                fm = ma.forecast_month
                if hasattr(fm, 'date'):
                    fm = fm.date()
                monthly_assumptions_raw[proj.id][fm.isoformat()] = {
                    "dev_cost_increase": ma.dev_cost_increase,
                    "lot_paydown_amount": ma.lot_paydown_amount,
                    "lots_released": ma.lots_released
                }

        lb_result = calculate_lb(lb_projects, monthly_assumptions_raw, num_months)

        for row_idx, proj in enumerate(lb_projects, start=2):
            ws_lb.cell(row=row_idx, column=1, value=proj.dept_code)
            ws_lb.cell(row=row_idx, column=2, value=proj.builder)
            ws_lb.cell(row=row_idx, column=3, value=proj.project_name)
            ws_lb.cell(row=row_idx, column=4, value=proj.phase)
            currency_cell(ws_lb, row_idx, 5, proj.current_balance)
            currency_cell(ws_lb, row_idx, 6, proj.total_budget)

            proj_data = lb_result["per_project"].get(proj.id, {})
            balances = proj_data.get("balances", [0.0] * num_months)
            for m_idx, bal in enumerate(balances):
                currency_cell(ws_lb, row_idx, 7 + m_idx, bal)

        # Totals row
        total_row = len(lb_projects) + 2
        ws_lb.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        totals = lb_result.get("totals", [0.0] * num_months)
        for m_idx, tot in enumerate(totals):
            currency_cell(ws_lb, total_row, 7 + m_idx, tot)

    auto_width(ws_lb)

    # ---- HHH Sheet ----
    ws_hhh = wb.create_sheet("HHH_OW_MMC")
    hhh_headers = ["Investment", "Current Balance", "Loan Amount", "Monthly Change", "Maturity"] + month_labels
    for col_idx, h in enumerate(hhh_headers, start=1):
        ws_hhh.cell(row=1, column=col_idx, value=h)
    style_header_row(ws_hhh, 1, len(hhh_headers))

    hhh_investments = db.query(models.HHHInvestment).filter(
        models.HHHInvestment.version_id == version_id
    ).all()

    if hhh_investments:
        hhh_result = calculate_hhh(hhh_investments, num_months)

        for row_idx, inv in enumerate(hhh_investments, start=2):
            ws_hhh.cell(row=row_idx, column=1, value=inv.investment_name)
            currency_cell(ws_hhh, row_idx, 2, inv.current_balance)
            currency_cell(ws_hhh, row_idx, 3, inv.loan_amount)
            monthly_change = inv.monthly_increase if inv.investment_name != "HH_TOM" else -inv.monthly_deduction
            currency_cell(ws_hhh, row_idx, 4, monthly_change)
            ws_hhh.cell(row=row_idx, column=5, value=inv.maturity_date)

            inv_data = hhh_result["per_investment"].get(inv.investment_name, {})
            balances = inv_data.get("balances", [0.0] * num_months)
            for m_idx, bal in enumerate(balances):
                currency_cell(ws_hhh, row_idx, 6 + m_idx, bal)

        # Totals row
        total_row = len(hhh_investments) + 2
        ws_hhh.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        totals = hhh_result.get("totals", [0.0] * num_months)
        for m_idx, tot in enumerate(totals):
            currency_cell(ws_hhh, total_row, 6 + m_idx, tot)

    auto_width(ws_hhh)

    # ---- Land Bucket (detail) ----
    ws_lb2 = wb.create_sheet("Land Bucket")
    ws_lb2.cell(row=1, column=1, value="Land Bucket Builder Rollup").font = Font(bold=True, size=14)
    if lb_projects and lb_result:
        per_builder = lb_result.get("per_builder", {})
        ws_lb2.cell(row=2, column=1, value="Builder")
        for m_idx, m in enumerate(month_labels):
            ws_lb2.cell(row=2, column=2 + m_idx, value=m)
        style_header_row(ws_lb2, 2, 1 + len(month_labels))

        row_idx = 3
        for builder, balances in per_builder.items():
            ws_lb2.cell(row=row_idx, column=1, value=builder)
            for m_idx, bal in enumerate(balances):
                currency_cell(ws_lb2, row_idx, 2 + m_idx, bal)
            row_idx += 1

    auto_width(ws_lb2)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"SNCC_Forecast_{version.label.replace(' ', '_')}_{today.isoformat()}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
