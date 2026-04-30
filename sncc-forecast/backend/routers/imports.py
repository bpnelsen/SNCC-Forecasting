from __future__ import annotations
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from datetime import date, datetime
from typing import Optional
from dateutil.relativedelta import relativedelta
import openpyxl
import io
import models
from database import get_db

router = APIRouter()


def map_loan_type(loan_program: str) -> str:
    if not loan_program:
        return "SFR"
    lp = loan_program.strip()
    if "Single Family Residential Construction" in lp:
        return "SFR"
    elif "Multifamily Residential Construction" in lp:
        return "MF"
    elif "Multiple Lots Lot Loan" in lp:
        return "AD"
    elif "Raw Land" in lp:
        return "RawLand"
    elif "Lot" in lp:
        return "FinishedLots"
    else:
        return "SFR"


def parse_date(value) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"]:
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_float(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        val = value.strip().replace("$", "").replace(",", "")
        try:
            return float(val)
        except ValueError:
            return None
    return None


@router.post("/import/current-report")
async def import_current_report(
    version_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        version = db.query(models.ForecastVersion).filter(models.ForecastVersion.id == version_id).first()
        if not version:
            raise HTTPException(status_code=404, detail="Version not found")

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        # Find the Current Report sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "Current Report" in name or "current report" in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            # Try first sheet
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Get headers from first row
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value is not None:
                headers[str(cell.value).strip()] = col_idx

        def get_col(row, header_name):
            col = headers.get(header_name)
            if col:
                return row[col - 1].value
            return None

        today = date.today()

        # Delete existing loans for this version
        db.query(models.Loan).filter(models.Loan.version_id == version_id).delete()

        imported = 0
        flagged_loans = []

        for row in ws.iter_rows(min_row=2, values_only=False):
            # Skip empty rows
            first_val = row[0].value if row else None
            if first_val is None:
                continue

            # Extract values
            loan_number = get_col_val(row, headers, "Loan Number") or get_col_val(row, headers, "Loan #") or get_col_val(row, headers, "Loan No")
            borrower = get_col_val(row, headers, "Borrower") or get_col_val(row, headers, "Borrower Name")
            loan_program = get_col_val(row, headers, "Loan Program") or get_col_val(row, headers, "Product Type")
            loan_amount_val = parse_float(get_col_val(row, headers, "Loan Amount") or get_col_val(row, headers, "Original Loan Amount"))
            funded_date_val = parse_date(get_col_val(row, headers, "Funded Date") or get_col_val(row, headers, "Fund Date") or get_col_val(row, headers, "Origination Date"))
            maturity_date_val = parse_date(get_col_val(row, headers, "Maturity Date") or get_col_val(row, headers, "Maturity"))
            current_balance_val = parse_float(get_col_val(row, headers, "Current Balance") or get_col_val(row, headers, "Outstanding Balance") or get_col_val(row, headers, "Balance"))
            interest_rate_val = parse_float(get_col_val(row, headers, "Interest Rate") or get_col_val(row, headers, "Rate") or get_col_val(row, headers, "Note Rate"))
            development_name = get_col_val(row, headers, "Development Name") or get_col_val(row, headers, "Development") or get_col_val(row, headers, "Project Name")
            subdivision_name = get_col_val(row, headers, "Subdivision Name") or get_col_val(row, headers, "Subdivision")

            # Skip rows without meaningful data
            if not loan_number and not borrower:
                continue

            loan_type = map_loan_type(str(loan_program) if loan_program else "")

            # Calculate loan_remaining
            loan_remaining = None
            if loan_amount_val is not None and current_balance_val is not None:
                loan_remaining = loan_amount_val - current_balance_val
            elif loan_amount_val is not None:
                loan_remaining = loan_amount_val

            projected_balance = loan_amount_val

            # Calculate months_remaining
            months_remaining = None
            if maturity_date_val:
                rd = relativedelta(maturity_date_val, today)
                months_remaining = rd.years * 12 + rd.months
                if maturity_date_val < today:
                    months_remaining = -1

            # Calculate monthly_draw_increase
            monthly_draw_increase = None
            if loan_remaining is not None and months_remaining is not None and months_remaining > 0:
                monthly_draw_increase = loan_remaining / months_remaining
            elif loan_remaining is not None and months_remaining is not None and months_remaining <= 0:
                monthly_draw_increase = 0.0

            # Interest rate as decimal if given as percentage
            if interest_rate_val is not None and interest_rate_val > 1:
                interest_rate_val = interest_rate_val / 100.0

            # Flag logic
            draw_flagged = False
            flag_reasons = []

            if months_remaining is not None and months_remaining <= 0:
                draw_flagged = True
                flag_reasons.append("Maturity date has passed")

            if monthly_draw_increase is not None and monthly_draw_increase < 0:
                draw_flagged = True
                flag_reasons.append("Negative draw - check loan type mapping")

            if (loan_remaining is not None and monthly_draw_increase is not None and
                    loan_remaining > 0 and monthly_draw_increase / loan_remaining > 0.25):
                draw_flagged = True
                flag_reasons.append("Unusually fast draw pace")

            if (maturity_date_val and loan_amount_val and loan_remaining is not None and
                    months_remaining is not None and 0 < months_remaining <= 2 and
                    loan_remaining > loan_amount_val * 0.20):
                draw_flagged = True
                flag_reasons.append("Near maturity with large remaining")

            draw_flag_reason = "; ".join(flag_reasons) if flag_reasons else None

            loan_obj = models.Loan(
                version_id=version_id,
                loan_number=str(loan_number) if loan_number else None,
                borrower=str(borrower) if borrower else None,
                loan_program=str(loan_program) if loan_program else None,
                loan_type=loan_type,
                loan_amount=loan_amount_val,
                funded_date=funded_date_val,
                maturity_date=maturity_date_val,
                current_balance=current_balance_val,
                loan_remaining=loan_remaining,
                interest_rate=interest_rate_val,
                development_name=str(development_name) if development_name else None,
                subdivision_name=str(subdivision_name) if subdivision_name else None,
                projected_balance=projected_balance,
                monthly_draw_increase=monthly_draw_increase,
                draw_flagged=draw_flagged,
                draw_flag_reason=draw_flag_reason
            )
            db.add(loan_obj)
            imported += 1

            if draw_flagged:
                flagged_loans.append({
                    "loan_number": str(loan_number) if loan_number else None,
                    "borrower": str(borrower) if borrower else None,
                    "flag_reason": draw_flag_reason,
                    "monthly_draw_increase": monthly_draw_increase,
                    "loan_type": loan_type
                })

        db.commit()

        # Update version import_date
        version.import_date = datetime.utcnow()
        db.commit()

        return {
            "imported": imported,
            "flagged": len(flagged_loans),
            "flagged_loans": flagged_loans
        }

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")


def get_col_val(row, headers, name):
    col = headers.get(name)
    if col:
        cell = row[col - 1]
        return cell.value
    return None


@router.post("/import/land-bucket")
async def import_land_bucket(
    version_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        version = db.query(models.ForecastVersion).filter(models.ForecastVersion.id == version_id).first()
        if not version:
            raise HTTPException(status_code=404, detail="Version not found")

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        # Read LB Data sheet
        lb_data_sheet = None
        lb_status_sheet = None
        for name in wb.sheetnames:
            if "LB Data" in name or "lb data" in name.lower():
                lb_data_sheet = wb[name]
            elif "LB Status" in name or "lb status" in name.lower():
                lb_status_sheet = wb[name]

        if not lb_data_sheet:
            lb_data_sheet = wb[wb.sheetnames[0]]
        if not lb_status_sheet and len(wb.sheetnames) > 1:
            lb_status_sheet = wb[wb.sheetnames[1]]

        # Read LB Status (current balances)
        status_map = {}  # dept_code -> current_balance
        if lb_status_sheet:
            status_headers = {}
            for col_idx, cell in enumerate(lb_status_sheet[1], start=1):
                if cell.value is not None:
                    status_headers[str(cell.value).strip()] = col_idx

            for row in lb_status_sheet.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                dept_col = status_headers.get("Dept Code") or status_headers.get("Department Code") or 1
                name_col = status_headers.get("Project Name") or status_headers.get("Name") or 2
                balance_col = status_headers.get("Total Amount") or status_headers.get("Balance") or 3
                dept_code = row[dept_col - 1]
                proj_name = row[name_col - 1]
                balance = parse_float(row[balance_col - 1])
                if dept_code:
                    status_map[str(dept_code).strip()] = {
                        "balance": balance,
                        "project_name": str(proj_name) if proj_name else None
                    }

        # Read LB Data
        data_headers = {}
        for col_idx, cell in enumerate(lb_data_sheet[1], start=1):
            if cell.value is not None:
                data_headers[str(cell.value).strip()] = col_idx

        # Delete existing
        db.query(models.LBProject).filter(models.LBProject.version_id == version_id).delete()

        imported = 0
        for row in lb_data_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            dept_code = get_row_val(row, data_headers, "Dept Code") or get_row_val(row, data_headers, "Department Code")
            builder = get_row_val(row, data_headers, "Builder") or get_row_val(row, data_headers, "Builder Name")
            project_code = get_row_val(row, data_headers, "Project Code") or get_row_val(row, data_headers, "Project")
            phase = get_row_val(row, data_headers, "Phase")
            total_lots = get_row_val(row, data_headers, "Lots") or get_row_val(row, data_headers, "Total Lots")
            total_budget = parse_float(get_row_val(row, data_headers, "Total Budget") or get_row_val(row, data_headers, "Budget"))
            dev_costs = parse_float(get_row_val(row, data_headers, "Development Costs") or get_row_val(row, data_headers, "Dev Costs") or get_row_val(row, data_headers, "Development Costs Remaining"))

            if not dept_code and not builder:
                continue

            dept_str = str(dept_code).strip() if dept_code else None
            status_info = status_map.get(dept_str, {}) if dept_str else {}
            current_balance = status_info.get("balance", 0.0)
            project_name = status_info.get("project_name") or str(project_code) if project_code else None

            proj = models.LBProject(
                version_id=version_id,
                dept_code=dept_str,
                builder=str(builder) if builder else None,
                project_name=project_name,
                phase=str(phase) if phase else None,
                total_lots=int(total_lots) if total_lots else None,
                current_balance=current_balance,
                total_budget=total_budget,
                development_costs_remaining=dev_costs
            )
            db.add(proj)
            imported += 1

        db.commit()
        version.import_date = datetime.utcnow()
        db.commit()

        return {"imported": imported}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"Land bucket import failed: {str(e)}")


def get_row_val(row, headers, name):
    col = headers.get(name)
    if col and col <= len(row):
        return row[col - 1]
    return None


@router.post("/import/hhh")
async def import_hhh(
    version_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        version = db.query(models.ForecastVersion).filter(models.ForecastVersion.id == version_id).first()
        if not version:
            raise HTTPException(status_code=404, detail="Version not found")

        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

        # Find HHH_OW_MMC sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "HHH" in name.upper() or "OW" in name.upper() or "MMC" in name.upper():
                sheet_name = name
                break
        if not sheet_name:
            sheet_name = wb.sheetnames[0]

        ws = wb[sheet_name]

        # Delete existing
        db.query(models.HHHInvestment).filter(models.HHHInvestment.version_id == version_id).delete()

        investment_names = ["HH_TOM", "HHH_OW", "HH-JKSN"]
        imported = 0

        # Scan rows for investment names
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue

            for cell_val in row:
                if cell_val and str(cell_val).strip().upper() in [n.upper() for n in investment_names]:
                    inv_name = str(cell_val).strip()
                    row_list = list(row)

                    # Extract values from the row - flexible column positions
                    values = [v for v in row_list if v is not None]

                    current_balance = None
                    loan_amount = None
                    monthly_increase = None
                    release_price = None
                    units_per_month = None
                    monthly_deduction = None
                    maturity_date_val = None

                    # Try to parse numeric values in order
                    float_vals = []
                    for v in row_list:
                        if v is None:
                            continue
                        if isinstance(v, (int, float)) and not isinstance(v, bool):
                            float_vals.append(float(v))
                        elif isinstance(v, str) and v.strip() not in investment_names:
                            parsed = parse_float(v)
                            if parsed is not None:
                                float_vals.append(parsed)
                            else:
                                d = parse_date(v)
                                if d:
                                    maturity_date_val = d

                    # Assign positionally
                    if len(float_vals) > 0:
                        current_balance = float_vals[0]
                    if len(float_vals) > 1:
                        loan_amount = float_vals[1]
                    if len(float_vals) > 2:
                        monthly_increase = float_vals[2]
                    if len(float_vals) > 3:
                        release_price = float_vals[3]
                    if len(float_vals) > 4:
                        units_per_month = int(float_vals[4])
                    if len(float_vals) > 5:
                        monthly_deduction = float_vals[5]

                    # Check for date in row
                    for v in row_list:
                        if isinstance(v, (date, datetime)):
                            maturity_date_val = v if isinstance(v, date) else v.date()

                    inv = models.HHHInvestment(
                        version_id=version_id,
                        investment_name=inv_name,
                        current_balance=current_balance,
                        loan_amount=loan_amount,
                        monthly_increase=monthly_increase or 0.0,
                        release_price=release_price,
                        units_per_month=units_per_month or 0,
                        monthly_deduction=monthly_deduction or 0.0,
                        maturity_date=maturity_date_val
                    )
                    db.add(inv)
                    imported += 1
                    break

        # If no investments found via row scan, create default empty ones
        if imported == 0:
            for inv_name in investment_names:
                inv = models.HHHInvestment(
                    version_id=version_id,
                    investment_name=inv_name,
                    current_balance=0.0,
                    loan_amount=0.0,
                    monthly_increase=0.0,
                    release_price=0.0,
                    units_per_month=0,
                    monthly_deduction=0.0,
                    maturity_date=None
                )
                db.add(inv)
                imported += 1

        db.commit()
        version.import_date = datetime.utcnow()
        db.commit()

        return {"imported": imported}

    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"HHH import failed: {str(e)}")
